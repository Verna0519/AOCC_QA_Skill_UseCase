#!/usr/bin/env python3
"""
AOCCQA-case-exporter

Fill the AOCC QA xlsx template from:
  (1) a Jira ticket's fields  -> Report sheet + filename
  (2) the previous agent's test cases (7-col standard) -> Test case sheet

Deterministic formatting only. No content judgement, filtering, or rewriting.
Bug list / Screenshot sheets and all formulas are left untouched.
"""

import argparse
import copy
import json
import os
import re
import sys
from datetime import datetime

import openpyxl

MAX_CASES = 200          # template rows 2..201
FIRST_DATA_ROW = 2

# Report sheet: label cell (col A) -> input cell (col C)
REPORT_CELLS = {
    "summary": "C2",         # Project cell <- full Jira Summary (tags kept)
    "test_date": "C3",
    "test_version": "C4",
    "tester": "C5",
    "link": "C6",            # New feature & Release Note
    "mcc": "C13",            # Test Country
    "test_environment": "C14",
}

# Test case sheet: previous-agent field -> template column letter
CASE_COLS = {
    "id": "A",
    "category": "E",
    "pre_condition": "F",
    "test_case": "G",
    "steps": "H",
    "expected_result": "I",
}

ILLEGAL_FILENAME = re.compile(r'[\\/:*?"<>|]')
UAT_QA_TAG = re.compile(r'\[UAT-QA\]', flags=re.IGNORECASE)
LEADING_TAG = re.compile(r'^\s*\[([^\]]+)\]\s*')   # e.g. "[EU] " -> "EU_"


def clean_summary_to_filename(summary: str, today: str) -> str:
    """
    Turn a Jira Summary into the export filename.

      "[UAT-QA][EU] Customized Bundle ... enhancement"
        -> remove [UAT-QA]        -> "[EU] Customized Bundle ... enhancement"
        -> "[EU] " becomes "EU_"  -> "EU_Customized Bundle ... enhancement"
        -> append suffix          -> "EU_Customized Bundle ... enhancement_Test Case_YYYYMMDD.xlsx"
    """
    base = UAT_QA_TAG.sub("", summary or "").strip()
    base = LEADING_TAG.sub(lambda m: m.group(1).strip() + "_", base).strip()
    if not base:
        base = "Untitled"
    name = f"{base}_Test Case_{today}.xlsx"
    name = ILLEGAL_FILENAME.sub("_", name)   # keep spaces/underscores as-is
    return name


def write_report(ws, jira: dict):
    """
    Fill Report input cells from Jira. Dynamic fields are filled only when a
    non-empty value is supplied; missing ones are left blank (never guessed).
    Returns a capture report listing captured vs blank for each dynamic field.
    """
    # dynamic Jira-sourced fields: key -> (cell, human label)
    dynamic = {
        "summary":          (REPORT_CELLS["summary"],          "Project (Summary)"),
        "test_date":        (REPORT_CELLS["test_date"],        "Test date"),
        "link":             (REPORT_CELLS["link"],             "New feature & Release Note (link)"),
        "mcc":              (REPORT_CELLS["mcc"],              "Test Country (MCC#)"),
        "test_environment": (REPORT_CELLS["test_environment"], "Test Environment"),
        "test_version":     (REPORT_CELLS["test_version"],     "Test Version"),
    }

    captured, blank = [], []
    for key, (cell, label) in dynamic.items():
        raw = jira.get(key)
        val = raw.strip() if isinstance(raw, str) else raw
        if val in (None, ""):
            blank.append({"field": label, "cell": cell})
        else:
            ws[cell] = val
            captured.append({"field": label, "cell": cell, "value": val})

    # Tester is derived, not raw: AOCCQA_<Assignee>. Blank assignee -> keep template default.
    assignee = (jira.get("assignee") or "").strip()
    tester_cell = REPORT_CELLS["tester"]
    if assignee:
        ws[tester_cell] = f"AOCCQA_{assignee}"
        captured.append({"field": "Tester", "cell": tester_cell, "value": f"AOCCQA_{assignee}"})
    else:
        blank.append({"field": "Tester (no Assignee on ticket)", "cell": tester_cell})

    return captured, blank


def write_cases(ws, cases: list):
    """Write test cases into rows 2..201. Feature is intentionally dropped."""
    for i, case in enumerate(cases):
        row = FIRST_DATA_ROW + i
        cid = str(case.get("id") or (i + 1)).strip()
        ws[f"{CASE_COLS['id']}{row}"] = cid
        ws[f"{CASE_COLS['category']}{row}"] = case.get("category", "")
        ws[f"{CASE_COLS['pre_condition']}{row}"] = case.get("pre_condition", "")
        ws[f"{CASE_COLS['test_case']}{row}"] = case.get("test_case", "")
        ws[f"{CASE_COLS['steps']}{row}"] = case.get("steps", "")
        ws[f"{CASE_COLS['expected_result']}{row}"] = case.get("expected_result", "")
        # B/C/D (platform), J/K/L (execution) left blank on purpose.


def main():
    ap = argparse.ArgumentParser(description="AOCCQA case exporter")
    ap.add_argument("--template", required=True, help="path to Test_Case_Template_Claude.xlsx")
    ap.add_argument("--input", required=True, help="path to input.json (jira + test_cases)")
    ap.add_argument("--outdir", default="/mnt/user-data/outputs", help="output directory")
    ap.add_argument("--date", default=None, help="override date as YYYYMMDD (default: today)")
    args = ap.parse_args()

    if not os.path.exists(args.template):
        sys.exit(f"ERROR: template not found: {args.template}")
    if not os.path.exists(args.input):
        sys.exit(f"ERROR: input.json not found: {args.input}")

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    jira = data.get("jira", {})
    cases = data.get("test_cases", [])

    if not cases:
        sys.exit("ERROR: 0 test cases supplied — nothing to export.")
    if len(cases) > MAX_CASES:
        sys.exit(f"ERROR: {len(cases)} cases exceed template limit of {MAX_CASES}.")

    today = args.date or datetime.now().strftime("%Y%m%d")

    wb = openpyxl.load_workbook(args.template)
    report_ws = wb["Report"]
    case_ws = wb["Test case"]

    captured, blanks = write_report(report_ws, jira)
    write_cases(case_ws, cases)
    # Bug list / Screenshot untouched by design.

    fname = clean_summary_to_filename(jira.get("summary", ""), today)
    os.makedirs(args.outdir, exist_ok=True)
    out_path = os.path.join(args.outdir, fname)
    wb.save(out_path)

    result = {
        "output_path": out_path,
        "filename": fname,
        "case_count": len(cases),
        "report_captured": captured,
        "report_blank": blanks,
        "sheets_preserved": ["Bug list", "Screenshot"],
        "formulas_touched": False,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
