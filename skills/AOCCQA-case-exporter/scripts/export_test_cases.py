#!/usr/bin/env python3
"""
AOCCQA-case-exporter

Fill the AOCC QA xlsx template from:
  (1) a Jira ticket's fields  -> Report sheet + filename
  (2) the previous agent's test cases (7-col standard) -> Test case sheet

Deterministic formatting only. No content judgement, filtering, or rewriting.
Bug list / Screenshot sheets and all formulas are left untouched.
"""

import argparse
import copy
import json
import os
import re
import sys
from datetime import datetime

import openpyxl

MAX_CASES = 200          # template rows 2..201
FIRST_DATA_ROW = 2

# Report sheet: label cell (col A) -> input cell (col C)
REPORT_CELLS = {
    "summary": "C2",         # Project cell <- full Jira Summary (tags kept)
    "test_date": "C3",
    "test_version": "C4",
    "tester": "C5",
    "link": "C6",            # New feature & Release Note
    "mcc": "C13",            # Test Country
    "test_environment": "C14",
}

# Test case sheet: previous-agent field -> template column letter
CASE_COLS = {
    "id": "A",
    "category": "E",
    "pre_condition": "F",
    "test_case": "G",
    "steps": "H",
    "expected_result": "I",
}

ILLEGAL_FILENAME = re.compile(r'[\\/:*?"<>|]')
UAT_QA_TAG = re.compile(r'\[UAT-QA\]', flags=re.IGNORECASE)
LEADING_TAG = re.compile(r'^\s*\[([^\]]+)\]\s*')   # e.g. "[EU] " -> "EU_"


def clean_summary_to_filename(summary: str, today: str) -> str:
    """
    Turn a Jira Summary into the export filename.

      "[UAT-QA][EU] Customized Bundle ... enhancement"
        -> remove [UAT-QA]        -> "[EU] Customized Bundle ... enhancement"
        -> "[EU] " becomes "EU_"  -> "EU_Customized Bundle ... enhancement"
        -> append suffix          -> "EU_Customized Bundle ... enhancement_Test Case_YYYYMMDD.xlsx"
    """
    base = UAT_QA_TAG.sub("", summary or "").strip()
    base = LEADING_TAG.sub(lambda m: m.group(1).strip() + "_", base).strip()
    if not base:
        base = "Untitled"
    name = f"{base}_Test Case_{today}.xlsx"
    name = ILLEGAL_FILENAME.sub("_", name)   # keep spaces/underscores as-is
    return name


def write_report(ws, jira: dict):
    """Fill Report sheet input cells. Return list of blanks that were skipped."""
    blanks = []

    def put(key, value):
        cell = REPORT_CELLS[key]
        val = (value or "").strip() if isinstance(value, str) else value
        if val in (None, ""):
            blanks.append((key, cell))
            return
        ws[cell] = val

    put("summary", jira.get("summary"))          # Project cell <- full Summary (tags kept)
    put("test_date", jira.get("test_date"))       # expect clean range YYYY/MM/DD-YYYY/MM/DD
    put("test_version", jira.get("test_version"))

    assignee = (jira.get("assignee") or "").strip()
    ws[REPORT_CELLS["tester"]] = f"AOCCQA_{assignee}" if assignee else ws[REPORT_CELLS["tester"]].value

    put("link", jira.get("link"))
    put("mcc", jira.get("mcc"))
    put("test_environment", jira.get("test_environment"))
    return blanks


def write_cases(ws, cases: list):
    """Write test cases into rows 2..201. Feature is intentionally dropped."""
    for i, case in enumerate(cases):
        row = FIRST_DATA_ROW + i
        cid = str(case.get("id") or (i + 1)).strip()
        ws[f"{CASE_COLS['id']}{row}"] = cid
        ws[f"{CASE_COLS['category']}{row}"] = case.get("category", "")
        ws[f"{CASE_COLS['pre_condition']}{row}"] = case.get("pre_condition", "")
        ws[f"{CASE_COLS['test_case']}{row}"] = case.get("test_case", "")
        ws[f"{CASE_COLS['steps']}{row}"] = case.get("steps", "")
        ws[f"{CASE_COLS['expected_result']}{row}"] = case.get("expected_result", "")
        # B/C/D (platform), J/K/L (execution) left blank on purpose.


def main():
    ap = argparse.ArgumentParser(description="AOCCQA case exporter")
    ap.add_argument("--template", required=True, help="path to Test_Case_Template_Claude.xlsx")
    ap.add_argument("--input", required=True, help="path to input.json (jira + test_cases)")
    ap.add_argument("--outdir", default="/mnt/user-data/outputs", help="output directory")
    ap.add_argument("--date", default=None, help="override date as YYYYMMDD (default: today)")
    args = ap.parse_args()

    if not os.path.exists(args.template):
        sys.exit(f"ERROR: template not found: {args.template}")
    if not os.path.exists(args.input):
        sys.exit(f"ERROR: input.json not found: {args.input}")

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    jira = data.get("jira", {})
    cases = data.get("test_cases", [])

    if not cases:
        sys.exit("ERROR: 0 test cases supplied — nothing to export.")
    if len(cases) > MAX_CASES:
        sys.exit(f"ERROR: {len(cases)} cases exceed template limit of {MAX_CASES}.")

    today = args.date or datetime.now().strftime("%Y%m%d")

    wb = openpyxl.load_workbook(args.template)
    report_ws = wb["Report"]
    case_ws = wb["Test case"]

    blanks = write_report(report_ws, jira)
    write_cases(case_ws, cases)
    # Bug list / Screenshot untouched by design.

    fname = clean_summary_to_filename(jira.get("summary", ""), today)
    os.makedirs(args.outdir, exist_ok=True)
    out_path = os.path.join(args.outdir, fname)
    wb.save(out_path)

    result = {
        "output_path": out_path,
        "filename": fname,
        "case_count": len(cases),
        "report_blanks": [{"field": k, "cell": c} for k, c in blanks],
        "sheets_preserved": ["Bug list", "Screenshot"],
        "formulas_touched": False,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
